import pandas as pd 
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side

# Define the input and output file paths
input_file_path = r'output\extracted_data_24-09-2024-23-38-41.txt'
output_file_path = r'output\extracted_data.xlsx'

# Initialize a list to hold the extracted data
data = []

# Read the input text file
with open(input_file_path, 'r', encoding='utf-8', errors='ignore') as file:
    lines = file.readlines()
    capture = False  # Flag to start capturing lines

    for line in lines:
        # Check for the start of the "show disks" section
        if line.strip() == "# show disks":
            capture = True  # Start capturing lines
            continue
        
        # Stop capturing when "Info:" is found
        if capture and "Info:" in line:
            break
        
        # Capture lines if within the relevant section
        if capture:
            # Process the line to extract relevant data
            parts = re.split(r'\s{2,}', line.strip())
            if len(parts) >= 5:  # Ensure there are enough columns
                location = parts[0].strip()
                serial_number = parts[1].strip()  # Serial Number
                rev = parts[3].strip()  # Rev
                data.append({'Location': location, 'Serial Number': serial_number, 'Rev': rev})

# Create a DataFrame from the extracted data
df = pd.DataFrame(data[1:])

# Attempt to write the DataFrame to an Excel file
try:
    # Check if the file already exists and prompt the user
    if os.path.exists(output_file_path):
        print(f"Warning: The file '{output_file_path}' already exists.")
        overwrite = input("Do you want to overwrite it? (y/n): ").strip().lower()
        if overwrite != 'y':
            print("Operation cancelled by the user.")
            exit()

    # Write the DataFrame to an Excel file
    df.to_excel(output_file_path, index=False, sheet_name='show_disk')

    # Load the workbook and select the active worksheet
    workbook = load_workbook(output_file_path)
    worksheet = workbook['show_disk']  # Select the sheet by name

    # Set the font for the entire worksheet to Courier
    courier_font = Font(name='Courier New')
    
    # Apply header style
    header_fill = PatternFill(start_color="FFFD00", end_color="FFFF00", fill_type="solid")  # Yellow fill for headers
    for cell in worksheet[1]:  # Header row
        cell.font = courier_font
        cell.fill = header_fill

    # Create a count of each Rev value
    rev_count = df['Rev'].value_counts()

    # Determine the minimum frequency for highlighting (for example, values that appear less than the most common one)
    min_count = rev_count.min()

    # Prepare to highlight values with less presence
    low_freq_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # Light red fill for low frequency values

    # Highlight low frequency Rev values in the worksheet
    for row in worksheet.iter_rows(min_row=2):  # Skip header row
        for cell in row:
            if cell.column_letter == 'C':  # Check the Rev column
                if rev_count[cell.value] == min_count:  # If the count matches the minimum frequency
                    cell.fill = low_freq_fill

    # Define border style
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    # Apply borders to the three columns
    for row in worksheet.iter_rows(min_row=1, max_col=3):  # Up to the third column
        for cell in row:
            cell.border = thin_border

    # Resize columns
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter  # Get the column letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)  # Add a little extra space
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # Save the changes to the workbook
    workbook.save(output_file_path)
    print(f'Data successfully extracted and saved to {output_file_path}')

except PermissionError:
    print(f"Error: Permission denied. Please close the Excel file if it is open or check your permissions for '{output_file_path}'.")
except Exception as e:
    print(f"An error occurred while saving the file: {e}")
